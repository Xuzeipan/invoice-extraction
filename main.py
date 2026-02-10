import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import re
import shutil
import platform

try:
    import PyPDF2
except ImportError:
    messagebox.showerror("缺少依赖", "请先安装PyPDF2：\npip install PyPDF2 openpyxl")
    raise

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, numbers
except ImportError:
    messagebox.showerror("缺少依赖", "请先安装openpyxl：\npip install openpyxl")
    raise


class InvoiceProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("数电发票信息提取工具")
        self.root.geometry("900x700")
        self.root.configure(bg='#f0f0f0')
        
        self.pdf_files = []
        self.excel_file = None
        
        # 预定义样式
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        self.setup_ui()
        
    def setup_ui(self):
        """设置界面布局"""
        title_label = tk.Label(
            self.root, 
            text="数电发票信息提取与重命名工具", 
            font=("Microsoft YaHei", 16, "bold"),
            bg='#f0f0f0',
            fg='#333333'
        )
        title_label.pack(pady=20)
        
        if platform.system() != 'Windows':
            tip_label = tk.Label(
                self.root,
                text=f"当前系统: {platform.system()}，请使用下方按钮选择文件",
                font=("Microsoft YaHei", 9),
                bg='#f0f0f0',
                fg='#FF6B6B'
            )
            tip_label.pack(pady=(0, 10))
        
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # PDF文件区域
        pdf_frame = tk.LabelFrame(
            main_frame, 
            text="PDF发票文件列表", 
            font=("Microsoft YaHei", 10),
            bg='#ffffff',
            fg='#333333',
            padx=10,
            pady=10
        )
        pdf_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.pdf_listbox = tk.Listbox(
            pdf_frame,
            font=("Consolas", 10),
            selectmode=tk.MULTIPLE,
            height=10
        )
        self.pdf_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        pdf_scrollbar = ttk.Scrollbar(pdf_frame, orient=tk.VERTICAL, command=self.pdf_listbox.yview)
        pdf_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.pdf_listbox.config(yscrollcommand=pdf_scrollbar.set)
        
        pdf_btn_frame = tk.Frame(main_frame, bg='#f0f0f0')
        pdf_btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(
            pdf_btn_frame, 
            text="添加PDF文件", 
            command=self.add_pdf_files
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            pdf_btn_frame, 
            text="清空列表", 
            command=self.clear_pdf_list
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            pdf_btn_frame, 
            text="移除选中", 
            command=self.remove_selected_pdf
        ).pack(side=tk.LEFT, padx=5)
        
        # Excel文件区域
        excel_frame = tk.LabelFrame(
            main_frame, 
            text="Excel模板文件", 
            font=("Microsoft YaHei", 10),
            bg='#ffffff',
            fg='#333333',
            padx=10,
            pady=10
        )
        excel_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.excel_label = tk.Label(
            excel_frame,
            text="尚未选择Excel文件",
            font=("Microsoft YaHei", 10),
            bg='#ffffff',
            fg='#999999',
            anchor='w'
        )
        self.excel_label.pack(fill=tk.X, pady=10)
        
        ttk.Button(
            excel_frame, 
            text="选择Excel模板", 
            command=self.select_excel_file
        ).pack(anchor='w', pady=(5, 0))
        
        # 操作按钮
        action_frame = tk.Frame(main_frame, bg='#f0f0f0')
        action_frame.pack(fill=tk.X, pady=20)
        
        self.process_btn = tk.Button(
            action_frame,
            text="开始处理（提取+写入+重命名）",
            font=("Microsoft YaHei", 12, "bold"),
            bg='#4CAF50',
            fg='white',
            activebackground='#45a049',
            activeforeground='white',
            padx=30,
            pady=10,
            command=self.process_all
        )
        self.process_btn.pack()
        
        # 日志区域
        log_frame = tk.LabelFrame(
            main_frame, 
            text="处理日志", 
            font=("Microsoft YaHei", 10),
            bg='#ffffff',
            fg='#333333'
        )
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(
            log_frame,
            font=("Consolas", 9),
            height=8,
            bg='#fafafa',
            fg='#333333',
            wrap=tk.WORD
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        log_scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=log_scrollbar.set)
            
    def add_pdf_files(self):
        """通过按钮添加PDF文件"""
        files = filedialog.askopenfilenames(
            title="选择PDF发票文件",
            filetypes=[("PDF文件", "*.pdf")]
        )
        for file_path in files:
            if file_path not in self.pdf_files:
                self.pdf_files.append(file_path)
                self.pdf_listbox.insert(tk.END, os.path.basename(file_path))
                self.log(f"添加PDF: {os.path.basename(file_path)}")
                
    def clear_pdf_list(self):
        """清空PDF列表"""
        self.pdf_files.clear()
        self.pdf_listbox.delete(0, tk.END)
        self.log("已清空PDF列表")
        
    def remove_selected_pdf(self):
        """移除选中的PDF"""
        selection = self.pdf_listbox.curselection()
        for index in reversed(selection):
            removed_file = self.pdf_files.pop(index)
            self.pdf_listbox.delete(index)
            self.log(f"移除: {os.path.basename(removed_file)}")
        
    def select_excel_file(self):
        """选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel模板文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        if file_path:
            self.excel_file = file_path
            self.excel_label.config(
                text=file_path,
                fg='#333333'
            )
            self.log(f"选择Excel模板: {os.path.basename(file_path)}")
            
    def log(self, message):
        """添加日志"""
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def extract_drawer(self, text):
        """
        专门提取开票人，处理换行情况
        策略：找到"开票人"关键字后，往后找第一个符合人名特征的行（2-4个汉字）
        """
        lines = text.split('\n')
        
        for i, line in enumerate(lines):
            if '开票人' in line:
                # 情况1：开票人和名字在同一行，如"开票人:高健铭"
                current = line.replace('开票人', '').replace(':', '').replace('：', '').strip()
                if current and self.is_valid_name(current):
                    return current
                
                # 情况2：名字在下一行
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    if self.is_valid_name(next_line):
                        return next_line
        
        return '高健铭'  # 默认值
    
    def is_valid_name(self, text):
        """
        判断文本是否像人名（2-4个汉字，不含数字、英文、特殊符号）
        """
        if not text:
            return False
        
        # 过滤掉明显不是名字的内容
        if any(x in text for x in ['¥', '公司', '电子', '发票', '号码', '2026', '9144']):
            return False
        
        # 匹配2-4个汉字（中文人名常见长度）
        return bool(re.match(r'^[\u4e00-\u9fa5]{2,4}$', text))
        
    def extract_invoice_data(self, pdf_path):
        """
        从PDF提取发票数据
        """
        try:
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = reader.pages[0].extract_text()
                
            if not text:
                self.log(f"  警告: {os.path.basename(pdf_path)} 无法提取文本（可能是扫描件）")
                return None
                
            data = {}
            
            # 1. 数电发票号码（20位数字）
            match = re.search(r'\b(\d{20})\b', text)
            data['invoice_no'] = match.group(1) if match else ''
            
            # 2. 开票日期 -> 转换为YYYY-MM-DD格式
            match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', text)
            if match:
                year, month, day = match.groups()
                data['date'] = f"{year}-{int(month):02d}-{int(day):02d}"
            else:
                data['date'] = ''
            
            # 3. 销方信息（通过税号定位）
            seller_match = re.search(
                r'91430100MA4TCG0Q2E\s*([\s\S]*?)\s*(91440300MA5H2BG470)', 
                text
            )
            if seller_match:
                seller_name = seller_match.group(1).replace('\n', '').strip()
                data['seller_name'] = seller_name
                data['seller_tax_no'] = '91440300MA5H2BG470'
            else:
                data['seller_name'] = '鼎越数科（深圳）信息技术有限公司'
                data['seller_tax_no'] = '91440300MA5H2BG470'
            
            # 4. 购方信息
            data['buyer_name'] = '湖南新飞创不良资产处置有限公司'
            data['buyer_tax_no'] = '91430100MA4TCG0Q2E'
            
            # 5. 金额提取策略
            amounts = re.findall(r'[¥￥]\s*([\d,]+\.\d{2})', text)
            if len(amounts) >= 3:
                nums = [(float(a.replace(',', '')), a) for a in amounts]
                nums.sort(key=lambda x: x[0], reverse=True)
                
                data['total'] = nums[0][1].replace(',', '')
                data['amount'] = nums[1][1].replace(',', '')
                data['tax'] = nums[2][1].replace(',', '')
            else:
                data['amount'] = ''
                data['tax'] = ''
                data['total'] = ''
            
            # 6. 开票人（使用专门的提取方法）
            data['drawer'] = self.extract_drawer(text)
            
            # 7. 货物或应税劳务名称
            item_match = re.search(r'(\*信息系统服务\*技术服务费?)', text)
            data['item_name'] = item_match.group(1) if item_match else '*信息系统服务*技术服务费'
            
            # 8. 备注（从文本末尾查找业务描述）
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            data['remark'] = ''
            
            for line in reversed(lines):
                if any(x in line for x in ['开票人', data['drawer'], '鼎越', '新飞创', '9144', '2026年']):
                    continue
                if '¥' in line or '电子发票' in line or '增值税专用发票' in line:
                    continue
                if '月' in line and ('费' in line or '服务' in line or '项目' in line):
                    data['remark'] = line
                    break
            
            # 9. 对应月份（从备注提取完整格式）
            if data['remark']:
                month_match = re.search(r'(\d{4}年\d{1,2}[-~]\d{1,2}月|\d{4}年\d{1,2}月)', data['remark'])
                if month_match:
                    month_str = month_match.group(1)
                    data['month'] = month_str.replace('月', '月份')
                else:
                    month_match = re.search(r'(\d{1,2}[-~]\d{1,2}月|\d{1,2}月)', data['remark'])
                    if month_match:
                        month_part = month_match.group(1).replace('月', '月份')
                        if data['date']:
                            year = data['date'][:4]
                            data['month'] = f"{year}年{month_part}"
                        else:
                            data['month'] = month_part
                    else:
                        data['month'] = ''
            else:
                data['month'] = ''
            
            return data
            
        except Exception as e:
            self.log(f"提取失败 {os.path.basename(pdf_path)}: {str(e)}")
            return None
            
    def find_total_row(self, ws):
        """
        查找合计行的位置（包含"合计"字样的行）
        返回行号（从1开始）
        """
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value  # 检查第一列
            if cell_value and '合计' in str(cell_value):
                return row
        return None
    
    def apply_cell_style(self, cell, is_number=False, is_amount=False):
        """
        应用单元格样式
        - is_amount: 是否为金额列（应用千分位格式）
        - is_number: 是否为数字列（居中对齐）
        """
        # 应用边框
        cell.border = self.thin_border
        
        # 应用对齐
        if is_number or is_amount:
            cell.alignment = self.center_alignment
        else:
            cell.alignment = self.left_alignment
        
        # 应用数字格式（千分位）
        if is_amount and cell.value:
            cell.number_format = '#,##0.00'
    
    def write_to_excel(self, excel_path, data_list):
        """
        将数据写入Excel，在合计行上方插入，不覆盖合计行，并应用样式
        """
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            total_row = self.find_total_row(ws)
            
            if total_row:
                self.log(f"  找到合计行在第{total_row}行")
                
                # 在合计行前插入足够的空行
                ws.insert_rows(total_row, len(data_list))
                self.log(f"  已在合计行前插入{len(data_list)}行")
                
                # 从新插入的第一行开始写入数据
                for idx, data in enumerate(data_list):
                    row = total_row + idx
                    self.write_row_data(ws, row, idx + 1, data)
                    self.log(f"  写入第{row}行: 发票{data['invoice_no'][:8]}... 开票人:{data['drawer']}")
            else:
                # 没找到合计行，从第2行开始写入
                self.log("  未找到合计行，从第2行开始写入")
                start_row = 2
                for idx, data in enumerate(data_list):
                    self.write_row_data(ws, start_row + idx, idx + 1, data)
            
            # 保存到新文件
            output_path = excel_path.replace('.xlsx', '_已填写.xlsx')
            counter = 1
            original_output = output_path
            while os.path.exists(output_path):
                name, ext = os.path.splitext(original_output)
                output_path = f"{name}_{counter}{ext}"
                counter += 1
                
            wb.save(output_path)
            return output_path
            
        except Exception as e:
            raise Exception(f"写入Excel失败: {str(e)}")
            
    def write_row_data(self, ws, row, seq_no, data):
        """写入一行数据到指定行，并应用样式"""
        # 定义金额列（第11、12、13列）
        amount_columns = {11, 12, 13}
        
        # 写入数据并应用样式
        columns_data = [
            (1, seq_no),  # 序号
            (2, ''),      # 发票代码
            (3, ''),      # 发票号码
            (4, data['invoice_no']),  # 数电发票号码
            (5, data['seller_tax_no']),  # 销方识别号
            (6, data['seller_name']),    # 销方名称
            (7, data['buyer_tax_no']),   # 购方识别号
            (8, data['buyer_name']),     # 购买方名称
            (9, data['date']),           # 开票日期
            (10, data['item_name']),     # 货物或应税劳务名称
            (11, float(data['amount']) if data['amount'] else ''),  # 金额
            (12, float(data['tax']) if data['tax'] else ''),        # 税额
            (13, float(data['total']) if data['total'] else ''),    # 价税合计
            (14, '电子发票服务平台'),     # 发票来源
            (15, '数电发票（增值税专用发票）'),  # 发票票种
            (16, '正常'),                # 发票状态
            (17, '是'),                  # 是否正数发票
            (18, '正常'),                # 发票风险等级
            (19, data['drawer']),        # 开票人
            (20, data['remark']),        # 备注
            (21, data['month']),         # 对应月份
            (22, '')                     # 项目名称
        ]
        
        for col, value in columns_data:
            cell = ws.cell(row=row, column=col, value=value)
            # 应用样式：金额列使用数字格式，其他列使用文本格式
            is_amount = col in amount_columns
            self.apply_cell_style(cell, is_number=(col in {1, 11, 12, 13}), is_amount=is_amount)
            
    def rename_pdf(self, pdf_path, data):
        """
        按规则重命名PDF文件
        规则：备注+金额+日期（无备注则为金额+日期）
        """
        try:
            remark = data.get('remark', '')
            total = data.get('total', '')
            date = data.get('date', '')
            
            def clean_filename(text):
                invalid_chars = '\\/:*?"<>|'
                for char in invalid_chars:
                    text = text.replace(char, '_')
                return text
            
            if remark:
                new_name = f"{remark}+{total}+{date}.pdf"
            else:
                new_name = f"{total}+{date}.pdf"
                
            new_name = clean_filename(new_name)
            
            dir_name = os.path.dirname(pdf_path)
            new_path = os.path.join(dir_name, new_name)
            
            counter = 1
            base_new_path = new_path
            while os.path.exists(new_path):
                name, ext = os.path.splitext(base_new_path)
                new_path = f"{name}_{counter}{ext}"
                counter += 1
            
            shutil.move(pdf_path, new_path)
            return new_path
            
        except Exception as e:
            raise Exception(f"重命名失败: {str(e)}")
            
    def process_all(self):
        """处理所有文件"""
        if not self.pdf_files:
            messagebox.showwarning("警告", "请先添加PDF发票文件！")
            return
            
        if not self.excel_file:
            messagebox.showwarning("警告", "请先选择Excel模板文件！")
            return
            
        self.process_btn.config(state='disabled', text='处理中...')
        self.log("=" * 50)
        self.log("开始处理...")
        
        try:
            data_list = []
            success_count = 0
            
            for pdf_path in self.pdf_files:
                self.log(f"正在提取: {os.path.basename(pdf_path)}...")
                data = self.extract_invoice_data(pdf_path)
                if data:
                    data_list.append(data)
                    success_count += 1
                    self.log(f"  ✓ 发票:{data['invoice_no'][:8]}... 金额:{data['total']} 开票人:{data['drawer']}")
                else:
                    self.log(f"  ✗ 提取失败")
            
            if not data_list:
                messagebox.showerror("错误", "未能从PDF中提取到有效数据！")
                return
                
            self.log(f"\n正在写入Excel...")
            output_excel = self.write_to_excel(self.excel_file, data_list)
            self.log(f"  ✓ Excel已保存: {os.path.basename(output_excel)}")
            
            self.log(f"\n正在重命名PDF文件...")
            for i, (pdf_path, data) in enumerate(zip(self.pdf_files, data_list)):
                try:
                    new_path = self.rename_pdf(pdf_path, data)
                    self.pdf_files[i] = new_path
                    self.log(f"  ✓ {os.path.basename(new_path)}")
                except Exception as e:
                    self.log(f"  ✗ 重命名失败: {str(e)}")
            
            self.log("\n" + "=" * 50)
            self.log(f"处理完成！成功提取{success_count}张发票")
            
            messagebox.showinfo(
                "完成", 
                f"处理完成！\n\n"
                f"成功提取: {success_count}张发票\n"
                f"Excel保存至: {os.path.basename(output_excel)}\n"
                f"已应用样式：边框、千分位数字格式、开票人提取"
            )
            
            self.clear_pdf_list()
            
        except Exception as e:
            messagebox.showerror("错误", f"处理过程中发生错误:\n{str(e)}")
            self.log(f"错误: {str(e)}")
            
        finally:
            self.process_btn.config(state='normal', text='开始处理（提取+写入+重命名）')


if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceProcessorApp(root)
    root.mainloop()